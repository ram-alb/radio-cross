from openpyxl import Workbook


def fill_headers(sheet):
    """
    Fill headers for radio cross report in excel.

    Args:
        sheet (openpyxl sheet obj): an active sheet
    """
    headers = [
        'Subnetwork',
        'Serial',
        'Radio',
        'Site1',
        'Site1 sector',
        'Site2',
        'Site2 sector',
    ]

    for column, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=column, value=header)


def fill_radio_params(sheet, row, radio_params):
    """
    Fill sitename and sector for radio into excel report.

    Args:
        sheet (openpyxl sheet obj): an active sheet
        row (int): a row number
        radio_params (dict): keys - sitename, values - sector used a radio
    """
    col = 4
    for parameter_name, parameter_value in radio_params.items():
        if parameter_name == 'subnetwork':
            sheet.cell(row=row, column=1, value=parameter_value)
        else:
            sheet.cell(row=row, column=col, value=parameter_name)
            col += 1
            sheet.cell(row=row, column=col, value=parameter_value)
            col += 1


def fill_report(radio_data):
    """
    Fill radio cross information to excel report.

    Args:
        radio_data (dict): keys - radio serials, values - dict with radio data

    Returns:
        str: a report path
    """
    work_book = Workbook()
    sheet = work_book.active

    fill_headers(sheet)

    row = 2
    for radio, radio_params in radio_data.items():
        serial, radio_type = radio.split(':')
        sheet.cell(row=row, column=2, value=serial)
        sheet.cell(row=row, column=3, value=radio_type)

        fill_radio_params(sheet, row, radio_params)

        row += 1

    report_path = 'reports/radio_cross.xlsx'
    work_book.save(report_path)
    return report_path
